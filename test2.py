# -*- coding:utf-8

import openpyxl
import sys
import re

inf=sys.argv[1].strip()
outf=sys.argv[2].strip()

workbook = openpyxl.load_workbook(inf)
sheet_names = workbook.sheetnames
sheet1 = workbook[sheet_names[1]]

nrows=sheet1.nrows #获得表的行数
ncols=sheet1.ncols #获得表的列数

new_data = []

for i in range(nrows): #遍历原始数据的每行
	row_info=sheet1.row_values(i) #读取一行的数据
    if row_info[0]=='Sample':
        row_info[1]='chr'
        row_info.insert(2,'POS')
        row_info.insert(3,'type')
        row_info.insert(16,'dgv')
    else:
	    chr = row_info[1].split(':')[0]
        typ = row_info[1].split('_')[-1]
        pos = row_info[1].split(':')[1].split('_')[0]
        dgv = row_info[13].split('，')[2].replace('占比：','')
        row_info[1]=chr
        row_info.insert(2,pos)
        row_info.insert(3,typ)
        row_info.insert(16,dgv)
    new_data.append(row_info)
wb = openpyxl.Workbook()
ws = workbook.active
ws.title = "sheet1"
flg = 0
for lines in new_data:
    flg +=1
    for i in range(len(lines)):
        ws.cell(flg,i+1,lines[i])
workbook.save(outf)